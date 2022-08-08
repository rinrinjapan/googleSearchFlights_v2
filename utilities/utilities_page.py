import csv
import inspect
import logging
import softest
from openpyxl import load_workbook


class Utilities(softest.TestCase):
    def custom_logger(logLevel = logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        fh = logging.FileHandler('automation.log', mode ='a')
        formatter = logging.Formatter('%(asctime)s - %(levelname)s -%(name)s : %(message)s', datefmt='%M:%D:%Y %I:%M:%S %p')
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger


    def assertion_value_flights(self, list_flights, value):
        for item in list_flights:
            self.soft_assert(self.assertEqual, item.text, value)
            if value == item.text:
                print("matched")
            else:
                print("not matched")
        self.assert_all()


    def read_excel_file( file_name, sheet):
        excel_do_vao = load_workbook(filename = file_name)
        sheet_do_vao = excel_do_vao[sheet]
        total_row = sheet_do_vao.max_row
        total_column = sheet_do_vao.max_column
        data_list = []
        for i in range(2, total_row + 1):
            row = []
            for j in range(1, total_column + 1):
                row.append(sheet_do_vao.cell(row=i, column=j).value)
            data_list.append(row)
        return data_list


    def read_file_csv(file_name):
        datalist = []
        csvdata = open(file_name,"r")
        reader = csv.reader(csvdata)
        next(reader)
        for row in reader:
            datalist.append(row)
        return datalist

